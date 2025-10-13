# pip install openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from pathlib import Path

S = 1.2  # cell side length
PI_2 = 1.57079632679
URI = "https://fuel.gazebosim.org/1.0/NGD1004/models/NIST maze wall 120"

def side_present(side) -> bool:
    """Treat a side as present if it has a style or a color."""
    if side is None:
        return False
    if getattr(side, "style", None):
        return True
    col = getattr(side, "color", None)
    return getattr(col, "rgb", None) is not None

def collect_canonical_borders(ws):
    """
    Return a set of (row, col, kind) with kind in {'left','top'}.

    Canonicalisation:
      - keep 'left' and 'top' on (r,c)
      - map 'right' of (r,c)  -> 'left' of (r, c+1)
      - map 'bottom' of (r,c) -> 'top'  of (r+1, c)
    """
    min_col, min_row, max_col, max_row = range_boundaries(ws.calculate_dimension())
    lines = set()

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            b = ws.cell(row=r, column=c).border
            if side_present(b.left):
                lines.add((r, c, "left"))
            if side_present(b.top):
                lines.add((r, c, "top"))
            if side_present(b.right):
                lines.add((r, c + 1, "left"))
            if side_present(b.bottom):
                lines.add((r + 1, c, "top"))
    return lines

def coords_of(r: int, c: int, kind: str):
    """
    Excel-like coordinates: origin at top-left of A1, y increases downward.
    CHANGE: For unrotated borders (theta=0, 'top'), subtract S from y.
      - 'top'  -> x=(c-1)*S, y=(r-1)*S - S, theta=0
      - 'left' -> x=(c-1)*S, y=(r-1)*S,     theta=pi/2
    """
    x_base = (c - 1) * S
    y_base = (r - 1) * S
    if kind == "top":
        return (x_base, y_base - S, 0.0)
    elif kind == "left":
        return (x_base, y_base, PI_2)
    else:
        raise ValueError("kind must be 'left' or 'top'")

def name_of(r: int, c: int, kind: str) -> str:
    return f"{get_column_letter(c)}{r}_{kind}"

def include_block(name: str, x: float, y: float, theta: float) -> str:
    # Gazebo pose: x y z roll pitch yaw ; z=0, roll=pitch=0, yaw=theta
    return (
f"""    <include>
      <name>{name}</name>
      <uri>{URI}</uri>
      <pose>{x:.10f} {y:.10f} 0 0 0 {theta:.11f}</pose>
    </include>"""
    )

def main(xlsx_path: str, sheet_name: str | None = None, out_path: str = "maze.txt"):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header = """<?xml version="1.0"?>
<sdf version="1.8">
  <model name="nist_square_120">
    <static>1</static>"""

    footer = """  </model>
</sdf>"""

    borders = collect_canonical_borders(ws)

    # Stable ordering: by row, then col, then left before top
    kind_order = {"left": 0, "top": 1}
    blocks = [header]
    for r, c, kind in sorted(borders, key=lambda t: (t[0], t[1], kind_order[t[2]])):
        name = name_of(r, c, kind)
        x, y, theta = coords_of(r, c, kind)
        blocks.append(include_block(name, x, y, theta))
    blocks.append(footer)

    # Write to maze.txt (UTF-8)
    Path(out_path).write_text("\n".join(blocks) + "\n", encoding="utf-8")
    print(f"Wrote {len(blocks)} border includes to {out_path}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python xlsx_borders_to_includes.py <xlsx-path> [sheet-name] [out-file]")
        sys.exit(1)
    xlsx = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    out = sys.argv[3] if len(sys.argv) > 3 else "maze.txt"
    main(xlsx, sheet, out)
